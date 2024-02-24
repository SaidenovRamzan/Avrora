import logging
import subprocess
from tkinter import messagebox
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas


def convert_xlsx_file_for_linux(xlsx_file, output_dir):
    try:
        command = [
            "libreoffice",
            "--headless",
            "--convert-to",
            "pdf",
            xlsx_file,
            "--outdir",
            output_dir,
        ]

        subprocess.run(command)
        logging.info(f"Conversion successful: {xlsx_file} -> {output_dir}")
    except Exception as e:
        logging.info(f"Error converting {xlsx_file} to PDF: {str(e)}")
        messagebox.showerror("Error", f"Error converting {xlsx_file} to PDF: {str(e)}")


def convert_xlsx_to_pdf_for_windows(xlsx_file, output_dir):
    try:
        wb = load_workbook(xlsx_file)
        worksheet = wb.active
        max_row = worksheet.max_row
        max_column = worksheet.max_column

        # Создаем PDF-файл
        c = canvas.Canvas(output_dir + "/result.pdf", pagesize=landscape(letter))

        top_margin = 3 * inch
        left_margin = right_margin = bottom_margin = 0.5 * inch
        cell_width = (15 * inch - left_margin - right_margin) / max_column
        cell_height = (8.5 * inch - top_margin - bottom_margin) / max_row

        # Получаем данные из листов Excel
        for row in range(1, max_row + 1):
            for column in range(1, max_column + 1):
                cell = worksheet.cell(row=row, column=column)
                text = str(cell.value) if str(cell.value) != "None" else " "

                x = left_margin + (column - 1) * cell_width
                y = 11 * inch - (top_margin + row * cell_height)

                c.drawString(x, y, text)
        c.save()
        logging.info(f"Conversion successful: {xlsx_file} -> {output_dir}")
    except Exception as e:
        logging.error(f"Error converting {xlsx_file} to PDF: {str(e)}")
        messagebox.showerror("Error", f"Error converting {xlsx_file} to PDF: {str(e)}")
